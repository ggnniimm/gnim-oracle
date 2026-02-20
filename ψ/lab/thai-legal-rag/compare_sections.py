"""
Compare วรรค counts: parser output vs Excel reference.

Usage:
    python compare_sections.py [--excel PATH] [--cache-dir PATH]

Defaults:
    --excel    ../sample-docs/จำนวนวรรคใน พ.ร.บ. และระเบียบ.xlsx
    --cache-dir  data/ocr_cache
"""
import argparse
import json
from pathlib import Path

import openpyxl


def load_excel(path: Path) -> dict[str, dict[int, int]]:
    """Return {sheet_type: {section_num: expected_paragraphs}}."""
    wb = openpyxl.load_workbook(path)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            sec, para = row[0], row[1]
            if sec is not None and para is not None:
                rows[int(sec)] = int(para)
        result[sheet_name] = rows
    return result


def load_cache(cache_dir: Path) -> list[dict]:
    """Return list of law dicts from law_*.json files."""
    laws = []
    for f in sorted(cache_dir.glob("law_*.json")):
        laws.append(json.loads(f.read_text()))
    return laws


def compare(law: dict, reference: dict[int, int], label: str) -> list[dict]:
    """Return list of diff rows."""
    diffs = []
    for sec in law["sections"]:
        sec_num = int(sec["number"])
        got = len(sec["paragraphs"])
        expected = reference.get(sec_num)
        if expected is None:
            diffs.append({
                "section": sec_num,
                "expected": "—",
                "got": got,
                "note": "ไม่มีใน Excel",
            })
        elif got != expected:
            diffs.append({
                "section": sec_num,
                "expected": expected,
                "got": got,
                "note": f"diff {got - expected:+d}",
            })
    # Check Excel sections not in parser
    parser_nums = {int(s["number"]) for s in law["sections"]}
    for sec_num, expected in sorted(reference.items()):
        if sec_num not in parser_nums:
            diffs.append({
                "section": sec_num,
                "expected": expected,
                "got": "—",
                "note": "ไม่มีใน parser",
            })
    return sorted(diffs, key=lambda d: d["section"] if isinstance(d["section"], int) else 0)


def print_report(label: str, law: dict, diffs: list[dict], total_sections: int) -> None:
    matched = total_sections - len(diffs)
    print(f"\n{'='*60}")
    print(f"  {label}  ({law['law_name'][:50]})")
    print(f"  Sections: {total_sections}  |  Match: {matched}  |  Diff: {len(diffs)}")
    print(f"{'='*60}")
    if not diffs:
        print("  ✓ ทุก section ตรงกับ Excel")
    else:
        print(f"  {'section':>8}  {'expected':>9}  {'got':>5}  note")
        print(f"  {'-'*8}  {'-'*9}  {'-'*5}  ----")
        for d in diffs:
            print(f"  {d['section']:>8}  {str(d['expected']):>9}  {str(d['got']):>5}  {d['note']}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="../sample-docs/จำนวนวรรคใน พ.ร.บ. และระเบียบ.xlsx")
    parser.add_argument("--cache-dir", default="data/ocr_cache")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    cache_dir = Path(args.cache_dir)

    print(f"Excel  : {excel_path}")
    print(f"Cache  : {cache_dir}")

    reference = load_excel(excel_path)
    laws = load_cache(cache_dir)

    # Map Excel sheet to law_type
    sheet_map = {
        "พ.ร.บ.": "พระราชบัญญัติ",
        "ระเบียบ": "ระเบียบ",
    }

    total_diffs = 0
    for sheet_name, law_type in sheet_map.items():
        ref = reference.get(sheet_name, {})
        law = next((l for l in laws if l["law_type"] == law_type), None)
        if law is None:
            print(f"\n[!] ไม่พบ cache สำหรับ {law_type}")
            continue
        diffs = compare(law, ref, sheet_name)
        print_report(sheet_name, law, diffs, len(law["sections"]))
        total_diffs += len(diffs)

    print(f"\n{'='*60}")
    print(f"  Total diffs: {total_diffs}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
