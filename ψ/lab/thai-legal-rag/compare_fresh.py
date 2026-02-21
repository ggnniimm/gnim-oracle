"""Compare fresh-parsed วรรค counts vs Excel reference (bypasses cache)."""
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, ".")
from src.ingestion.law_extractor import _split_paragraphs


def load_excel(path):
    wb = openpyxl.load_workbook(path)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec, para = row[0], row[1]
            if sec is not None and para is not None:
                rows[int(sec)] = int(para)
        result[sheet_name] = rows
    return result


reference = load_excel("../sample-docs/จำนวนวรรคใน พ.ร.บ. และระเบียบ.xlsx")

sheet_map = {
    "พ.ร.บ.": "law_98d856e0cc029ed2.json",
    "ระเบียบ": "law_6f2098489332d773.json",
}
total_diffs = 0

for sheet_name, fname in sheet_map.items():
    d = json.loads(Path(f"data/ocr_cache/{fname}").read_text())
    ref = reference[sheet_name]
    diffs = []
    for sec in d["sections"]:
        sec_num = int(sec["number"])
        fresh_paras = _split_paragraphs(sec["text"])
        got = len(fresh_paras)
        expected = ref.get(sec_num)
        if expected is not None and got != expected:
            diffs.append((sec_num, expected, got))

    matched = len(d["sections"]) - len(diffs)
    print(f"\n{sheet_name}: {len(d['sections'])} sections | Match {matched} | Diff {len(diffs)}")
    for sec_num, exp, got in diffs:
        print(f"  ข้อ/มาตรา {sec_num}: expected {exp}, got {got} ({got - exp:+d})")
    total_diffs += len(diffs)

print(f"\nTotal diffs (fresh parse): {total_diffs}")
